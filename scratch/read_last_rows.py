import openpyxl
import json

file_path = r'c:\Users\Annonymous\Documents\Fernando\Aplicaciones\PHP\ContaFC\database\Rep. archivo-final.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Archivo 2024']

max_row = sheet.max_row
data = []
for row in sheet.iter_rows(min_row=max_row - 1000, max_row=max_row, values_only=True):
    data.append(row)

print(json.dumps(data, indent=4, default=str))
