import openpyxl
import json

file_path = r'c:\Users\Annonymous\Documents\Fernando\Aplicaciones\PHP\ContaFC\database\Rep. archivo-final.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

target_codes = [
    312256640, 329033856, 873735810, 902001026, 903751811, 904962691, 
    915816577, 917027457, 918778242, 935555458, 949371009, 950581889, 
    952332674, 1002664322, 1021192323, 1037969539, 1050034305, 1138632835, 
    1222518915, 1239296131, 1256073347, 1289627779, 1457399939, 1490954371, 
    1558063235, 1574840451, 1641949315, 1965045891, 1977715841, 1980072322
]

# Convertir a strings para comparación flexible
target_str = [str(c) for c in target_codes]

results = []

for sheet_name in wb.sheetnames:
    print(f"Buscando en hoja: {sheet_name}...")
    sheet = wb[sheet_name]
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        for col_idx, cell_value in enumerate(row):
            if cell_value is not None and str(cell_value) in target_str:
                results.append({
                    "sheet": sheet_name,
                    "row": row_idx + 1,
                    "col": col_idx + 1,
                    "value": str(cell_value),
                    "full_row": row
                })
        if len(results) > 100: break # Limite de resultados

print(json.dumps(results, indent=4, default=str))
