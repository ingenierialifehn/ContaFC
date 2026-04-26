import openpyxl
import json

file_path = r'c:\Users\Annonymous\Documents\Fernando\Aplicaciones\PHP\ContaFC\database\Rep. archivo-final.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

# Nombres de terceros encontrados en la DB (batch 2026-04-05 11:54:24)
target_names = [
    "ESLY DALENY CRUZ GUERRERO",
    "LOURDES ANTONIA GAMEZ HERNANDEZ",
    "RIGOBERTO CANACA VARELA",
    "LETICIA ABIGAIL CANALES CANACA"
]

results = []

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        row_str = " ".join([str(c) for c in row if c is not None])
        for name in target_names:
            if name in row_str:
                results.append({
                    "name": name,
                    "sheet": sheet_name,
                    "row": row_idx + 1,
                    "full_row": row
                })
        if len(results) > 100: break

print(json.dumps(results, indent=4, default=str))
