import openpyxl

file_path = r'c:\Users\Annonymous\Documents\Fernando\Aplicaciones\PHP\ContaFC\database\Rep. archivo-final.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

target = "1002664322"

for name in wb.sheetnames:
    sheet = wb[name]
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        for c_idx, cell in enumerate(row):
            if str(cell) == target:
                print(f"ENCONTRADO en {name}, fila {r_idx+1}, col {c_idx+1}")
                print(row)
