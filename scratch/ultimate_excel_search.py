import openpyxl
import json
import re

file_path = r'c:\Users\Annonymous\Documents\Fernando\Aplicaciones\PHP\ContaFC\database\Rep. archivo-final.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

target_codes = [
    '312256640', '329033856', '873735810', '902001026', '903751811', '904962691', 
    '915816577', '917027457', '918778242', '935555458', '949371009', '950581889', 
    '952332674', '1002664322', '1021192323', '1037969539', '1050034305', '1138632835', 
    '1222518915', '1239296131', '1256073347', '1289627779', '1457399939', '1490954371', 
    '1558063235', '1574840451', '1641949315', '1965045891', '1977715841', '1980072322'
]

results = {}

for sheet_name in wb.sheetnames:
    print(f"Buscando en {sheet_name}...")
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        row_content = " ".join([str(c) for c in row if c is not None])
        # Limpiar caracteres no numéricos para búsqueda flexible
        clean_row = re.sub(r'\D', '', row_content)
        for code in target_codes:
            if code in clean_row:
                if code not in results:
                    # Intentar adivinar el nombre en la fila
                    # Basado en lo que vimos, suele estar en COMPANYcust (índice 5 en Archivo 2024)
                    # o en la columna 3 en Catalago (índice 2)
                    name = "Desconocido"
                    if sheet_name == 'Archivo 2024' or sheet_name == 'Archivo 2025':
                        name = row[5] if len(row) > 5 else "Desconocido"
                    elif sheet_name == 'Catalago':
                        name = row[2] if len(row) > 2 else "Desconocido"
                    else:
                        # Buscar cualquier string largo en la fila que no sea el código
                        for c in row:
                            if isinstance(c, str) and len(c) > 5 and not c.isdigit():
                                name = c
                                break
                    
                    results[code] = {
                        "name": name,
                        "sheet": sheet_name,
                        "row_data": row
                    }

print(json.dumps(results, indent=4, default=str))
