import openpyxl
import json

file_path = r'c:\Users\Annonymous\Documents\Fernando\Aplicaciones\PHP\ContaFC\database\Rep. archivo-final.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

if 'Catalago' not in wb.sheetnames:
    print(f"Error: Hoja 'Catalago' no encontrada.")
    exit(1)

sheet = wb['Catalago']

data = []
for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
    data.append(row)

print(json.dumps(data, indent=4, default=str))
