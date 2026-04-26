import openpyxl
import json

file_path = r'c:\Users\Annonymous\Documents\Fernando\Aplicaciones\PHP\ContaFC\database\Rep. archivo-final.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

target_name = "ESLY DALENY CRUZ GUERRERO"

results = []

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        row_str = " ".join([str(c) for c in row if c is not None])
        if target_name in row_str:
            results.append({
                "sheet": sheet_name,
                "row": row_idx + 1,
                "full_row": row
            })

print(json.dumps(results, indent=4, default=str))
